"""
Comprehensive analysis for VAF and HHOC - 2025 Annual + 2026 Q1 Reviews
"""
import openpyxl
from collections import defaultdict, Counter
from datetime import datetime
import json
import os

BASE = r"c:\vscode\annual-client-review\annual-client-review"

def safe_float(v):
    if v is None or v == 'NULL' or v == '':
        return 0.0
    try:
        return float(v)
    except:
        return 0.0

def categorize_ticket(title, notes):
    title = str(title).lower() if title else ''
    notes = str(notes).lower() if notes else ''
    text = title + ' ' + notes

    if any(w in text for w in ['patch', 'missing patch', 'windows update', ' kb']):
        return 'Patch Management'
    if any(w in text for w in ['crowdstrike', 'huntress', 'malware', 'threat', 'virus', 'security alert', 'falcon']):
        return 'Security & Endpoint'
    if any(w in text for w in ['backup', 'veeam', 'image backup', 'restore', 'recovery']):
        return 'Backup & DR'
    if any(w in text for w in ['monitor', 'alert', 'ops manager', 'manageengine', 'opsmanager', 'opsprb', 'probe', 'threshold']):
        return 'Monitoring & Alerts'
    if any(w in text for w in ['rmm', 'myremote', 'my remote', 'remote agent', 'connectwise', 'umbrella']):
        return 'RMM & Agent Management'
    if any(w in text for w in ['email', 'outlook', 'm365', 'office 365', 'mailbox', 'exchange', 'teams', 'sharepoint', 'onedrive', 'lightening', 'lightning']):
        return 'Email & M365'
    if any(w in text for w in ['firewall', 'sophos', 'vpn', 'network', 'switch', 'velocloud', 'vlan', 'dns', 'dhcp', 'cisco']):
        return 'Firewall & Network'
    if any(w in text for w in ['server', 'vm ', 'virtual machine', 'hyper-v', 'dc-', 'domain controller', 'active directory', 'gpo', 'group policy']):
        return 'Server Management'
    if any(w in text for w in ['new user', 'onboard', 'offboard', 'termination', 'new hire', 'setup user', 'disable user', 'archive user']):
        return 'User Onboarding/Offboarding'
    if any(w in text for w in ['password', 'reset password', 'account lock', 'mfa', '2fa', 'authenticator']):
        return 'Password & Account Management'
    if any(w in text for w in ['printer', 'print', 'scanner', 'scan']):
        return 'Printing & Scanning'
    if any(w in text for w in ['phone', 'voip', 'sip', 'did', 'call ', 'ring']):
        return 'Phone / VoIP'
    if any(w in text for w in ['install', 'software', 'application', 'upgrade', 'adobe', 'chrome']):
        return 'Software Installation & Updates'
    if any(w in text for w in ['workstation', 'laptop', 'desktop', 'hardware', 'pc ', 'computer']):
        return 'Workstation & Hardware'
    if any(w in text for w in ['file', 'folder', 'permission', 'share', 'drive']):
        return 'File & Permissions'
    if any(w in text for w in ['domain', 'ssl', 'certificate', 'dns record']):
        return 'Domain & SSL'
    if any(w in text for w in ['quickbook', 'sage', 'erp', 'dealertrack', 'routeone', 'deal pack']):
        return 'Application-Specific'
    if any(w in text for w in ['develop', 'coding', 'api', 'script', 'automation', 'website', 'web ']):
        return 'Software Development'
    return 'General IT Support'

PROACTIVE_CATS = {'Patch Management', 'Monitoring & Alerts', 'RMM & Agent Management', 'Security & Endpoint', 'Backup & DR'}

def analyze_time_entries(filepath, label):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    data = rows[1:]
    wb.close()

    result = {}
    result['total_entries'] = len(data)

    dates = [r[4] for r in data if r[4] and isinstance(r[4], datetime)]
    result['date_range'] = [min(dates).strftime('%Y-%m-%d'), max(dates).strftime('%Y-%m-%d')] if dates else ['N/A', 'N/A']

    tickets = set(r[1] for r in data if r[1])
    result['unique_tickets'] = len(tickets)

    total_nh = sum(safe_float(r[15]) for r in data)
    total_ah = sum(safe_float(r[16]) for r in data)
    total_onsite = sum(safe_float(r[17]) for r in data)
    total_hours = total_nh + total_ah + total_onsite
    total_drive = sum(safe_float(r[21]) for r in data)

    result['hours'] = {
        'nh': round(total_nh, 1),
        'ah': round(total_ah, 1),
        'onsite': round(total_onsite, 1),
        'total': round(total_hours, 1),
        'drive': round(total_drive, 1)
    }

    # By role type
    role_hours = defaultdict(lambda: {'nh': 0, 'ah': 0, 'onsite': 0, 'count': 0})
    for r in data:
        role = str(r[10]) if r[10] else 'Unknown'
        role_hours[role]['nh'] += safe_float(r[15])
        role_hours[role]['ah'] += safe_float(r[16])
        role_hours[role]['onsite'] += safe_float(r[17])
        role_hours[role]['count'] += 1
    result['by_role'] = {k: {kk: round(vv, 1) for kk, vv in v.items()} for k, v in role_hours.items()}

    # Technicians
    techs = defaultdict(float)
    for r in data:
        name = str(r[13]) if r[13] else 'Unknown'
        techs[name] += safe_float(r[15]) + safe_float(r[16]) + safe_float(r[17])
    result['technicians'] = {k: round(v, 1) for k, v in sorted(techs.items(), key=lambda x: -x[1])[:15]}
    result['tech_count'] = len(techs)

    # Monthly volume
    monthly = defaultdict(lambda: {'tickets': set(), 'nh': 0, 'ah': 0, 'onsite': 0, 'entries': 0})
    for r in data:
        if r[4] and isinstance(r[4], datetime):
            month = r[4].strftime('%Y-%m')
            monthly[month]['tickets'].add(r[1])
            monthly[month]['nh'] += safe_float(r[15])
            monthly[month]['ah'] += safe_float(r[16])
            monthly[month]['onsite'] += safe_float(r[17])
            monthly[month]['entries'] += 1
    result['monthly'] = {}
    for m in sorted(monthly.keys()):
        d = monthly[m]
        tk = len(d['tickets'])
        t = d['nh'] + d['ah'] + d['onsite']
        result['monthly'][m] = {
            'tickets': tk, 'nh': round(d['nh'], 1), 'ah': round(d['ah'], 1),
            'onsite': round(d['onsite'], 1), 'total': round(t, 1),
            'hrs_per_ticket': round(t / tk, 2) if tk > 0 else 0
        }

    # Weekend
    weekend = sum(1 for r in data if r[4] and isinstance(r[4], datetime) and r[4].weekday() >= 5)
    result['weekend_entries'] = weekend
    result['weekend_pct'] = round(weekend / max(len(data), 1) * 100, 1)

    # Contracts
    result['contracts'] = dict(Counter(str(r[12]) for r in data if r[12]))

    # Ticket categorization
    ticket_info = {}
    for r in data:
        tid = r[1]
        if tid not in ticket_info:
            ticket_info[tid] = {'title': r[2], 'notes': r[6], 'hours': 0}
        ticket_info[tid]['hours'] += safe_float(r[15]) + safe_float(r[16]) + safe_float(r[17])

    cat_stats = defaultdict(lambda: {'count': 0, 'hours': 0})
    for tid, info in ticket_info.items():
        cat = categorize_ticket(info['title'], info['notes'])
        cat_stats[cat]['count'] += 1
        cat_stats[cat]['hours'] += info['hours']

    result['categories'] = {k: {'count': v['count'], 'hours': round(v['hours'], 1),
                                 'avg': round(v['hours'] / v['count'], 2) if v['count'] > 0 else 0,
                                 'type': 'Proactive' if k in PROACTIVE_CATS else 'Reactive'}
                            for k, v in sorted(cat_stats.items(), key=lambda x: -x[1]['count'])}

    proactive = sum(v['count'] for k, v in result['categories'].items() if v['type'] == 'Proactive')
    reactive = sum(v['count'] for k, v in result['categories'].items() if v['type'] == 'Reactive')
    result['proactive_count'] = proactive
    result['reactive_count'] = reactive

    return result

def analyze_invoices(filepath, year=2025):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    data = rows[1:]
    wb.close()

    inv_year = [r for r in data if r[1] and isinstance(r[1], datetime) and r[1].year == year]

    result = {}
    result['total_line_items'] = len(inv_year)

    # By type
    type_rev = defaultdict(float)
    type_count = defaultdict(int)
    for r in inv_year:
        itype = str(r[2]) if r[2] else 'Unknown'
        type_rev[itype] += safe_float(r[9])
        type_count[itype] += 1
    result['by_type'] = {t: {'revenue': round(v, 2), 'count': type_count[t]}
                         for t, v in sorted(type_rev.items(), key=lambda x: -x[1])}
    result['total_revenue'] = round(sum(type_rev.values()), 2)

    # Monthly breakdown by category
    monthly_items = [r for r in inv_year if str(r[2]) == 'Monthly']
    monthly_cats = defaultdict(float)
    for r in monthly_items:
        item = str(r[3]) if r[3] else ''
        qb = str(r[8]) if r[8] else ''
        total = safe_float(r[9])

        if 'Labor' in qb or item.startswith('OffShore') or item.startswith('Tech_Support'):
            monthly_cats['Labor'] += total
        elif any(x in qb for x in ['Crowdstrike', 'Huntress', 'Malware']) or item in ['AVD','AVS','AVHS','AVMD','AVMS','AVMH']:
            monthly_cats['Security & Endpoint'] += total
        elif 'Backup' in qb or item in ['IB', 'VONE']:
            monthly_cats['Backup & Archiving'] += total
        elif any(x in qb for x in ['RMM', 'My Remote', 'My Ops', 'MyDisk']) or item in ['PMW','MR','OPS-BKP','OPS-NET','OPS-PRT','OPS-WF','MDU']:
            monthly_cats['Monitoring & Management'] += total
        elif 'SIP' in qb or item in ['SIP','DID','TFN','SMS','Incoming']:
            monthly_cats['VoIP / Telephony'] += total
        elif any(x in qb for x in ['VPS', 'Cloud']) or item in ['CL-GB','CL-VC','TB-BSTR','TB-PSTR','TB-RSTR']:
            monthly_cats['Cloud Hosting'] += total
        elif 'Secure Internet' in qb or item == 'SI':
            monthly_cats['Secure Internet'] += total
        elif 'PenTesting' in qb or item in ['PT','RTPT']:
            monthly_cats['Pen Testing'] += total
        elif 'Assessment' in qb or item in ['NA','SA']:
            monthly_cats['Assessment'] += total
        elif any(x in qb for x in ['Sophos', 'Edge', 'Velocloud']) or item in ['SO-1C4G','Edge-16M','VC-100M']:
            monthly_cats['Network & Firewall'] += total
        else:
            monthly_cats[f'Other ({item})'] += total

    result['monthly_categories'] = {k: round(v, 2) for k, v in sorted(monthly_cats.items(), key=lambda x: -x[1])}
    result['monthly_total'] = round(sum(monthly_cats.values()), 2)

    # Recurring
    recurring = [r for r in inv_year if str(r[2]) == 'Recurring']
    rec_items = defaultdict(float)
    for r in recurring:
        desc = str(r[4]) if r[4] else str(r[3])
        rec_items[desc] += safe_float(r[9])
    result['recurring'] = {k: round(v, 2) for k, v in sorted(rec_items.items(), key=lambda x: -x[1])}
    result['recurring_total'] = round(sum(rec_items.values()), 2)

    # WeeklyOut (overage)
    weeklyout = [r for r in inv_year if str(r[2]) == 'WeeklyOut']
    wo_items = []
    for r in weeklyout:
        wo_items.append({
            'date': r[1].strftime('%Y-%m-%d') if r[1] and isinstance(r[1], datetime) else '?',
            'item': str(r[3]),
            'desc': str(r[4])[:60] if r[4] else '',
            'qty': safe_float(r[6]),
            'price': safe_float(r[7]),
            'total': safe_float(r[9])
        })
    result['weeklyout'] = wo_items
    result['weeklyout_total'] = round(sum(safe_float(r[9]) for r in weeklyout), 2)

    # One-Time / Invoice
    onetime = [r for r in inv_year if str(r[2]) in ('One-Time', 'Invoice')]
    ot_items = []
    for r in onetime:
        ot_items.append({
            'date': r[1].strftime('%Y-%m-%d') if r[1] and isinstance(r[1], datetime) else '?',
            'item': str(r[3]),
            'desc': str(r[4])[:60] if r[4] else '',
            'qty': safe_float(r[6]),
            'price': safe_float(r[7]),
            'total': safe_float(r[9])
        })
    result['onetime'] = ot_items
    result['onetime_total'] = round(sum(i['total'] for i in ot_items), 2)

    # Weekly (contracted - usually $0 line total)
    weekly = [r for r in inv_year if str(r[2]) == 'Weekly']
    result['weekly_total'] = round(sum(safe_float(r[9]) for r in weekly), 2)

    # Monthly run rate
    monthly_by_month = defaultdict(float)
    for r in inv_year:
        if str(r[2]) == 'Monthly' and r[1] and isinstance(r[1], datetime):
            monthly_by_month[r[1].strftime('%Y-%m')] += safe_float(r[9])
    result['monthly_run_rate'] = {k: round(v, 2) for k, v in sorted(monthly_by_month.items())}

    # Labor detail by month for cycle detection
    labor_items = [r for r in inv_year if str(r[2]) == 'Monthly' and
                   ('Labor' in str(r[8]) or str(r[3]).startswith('OffShore') or str(r[3]).startswith('Tech_Support'))]
    labor_by_month = defaultdict(list)
    for r in labor_items:
        if r[1] and isinstance(r[1], datetime):
            labor_by_month[r[1].strftime('%Y-%m')].append({
                'item': str(r[3]),
                'desc': str(r[4])[:50] if r[4] else '',
                'qty': safe_float(r[6]),
                'price': safe_float(r[7]),
                'total': safe_float(r[9])
            })
    result['labor_by_month'] = dict(labor_by_month)

    # Latest month invoice detail (for representative breakdown)
    if monthly_by_month:
        latest = max(monthly_by_month.keys())
        latest_items = [r for r in inv_year if str(r[2]) == 'Monthly' and r[1] and
                       isinstance(r[1], datetime) and r[1].strftime('%Y-%m') == latest]
        result['latest_month'] = latest
        result['latest_invoice_detail'] = []
        for r in latest_items:
            result['latest_invoice_detail'].append({
                'item': str(r[3]),
                'desc': str(r[4])[:60] if r[4] else '',
                'qty': safe_float(r[6]),
                'price': safe_float(r[7]),
                'total': safe_float(r[9])
            })

    return result

# ── Run all analyses ──
print("Analyzing VAF 2025...")
vaf_2025_te = analyze_time_entries(os.path.join(BASE, 'clients/vaf/2025/vaf_tk_tm_2025.xlsx'), 'VAF 2025')
vaf_2025_inv = analyze_invoices(os.path.join(BASE, 'clients/vaf/2025/inv_from_2025_vaf.xlsx'), 2025)

print("Analyzing VAF 2026 Q1...")
vaf_2026_te = analyze_time_entries(os.path.join(BASE, 'clients/vaf/2026/vaf_tk_tm_2026.xlsx'), 'VAF 2026 Q1')

print("Analyzing HHOC 2025...")
hhoc_2025_te = analyze_time_entries(os.path.join(BASE, 'clients/hhoc/2025/hhoc_tk_te_2025.xlsx'), 'HHOC 2025')

print("Analyzing HHOC 2026 Q1...")
hhoc_2026_te = analyze_time_entries(os.path.join(BASE, 'clients/hhoc/2026/hhoc_tk_te_2026.xlsx'), 'HHOC 2026 Q1')

# Save results
results = {
    'vaf_2025_te': vaf_2025_te,
    'vaf_2025_inv': vaf_2025_inv,
    'vaf_2026_te': vaf_2026_te,
    'hhoc_2025_te': hhoc_2025_te,
    'hhoc_2026_te': hhoc_2026_te,
}

out_path = os.path.join(BASE, 'scripts/analysis_results.json')
with open(out_path, 'w') as f:
    json.dump(results, f, indent=2, default=str)

print(f"\nResults saved to {out_path}")

# Print summaries
for key, label in [('vaf_2025_te', 'VAF 2025'), ('vaf_2026_te', 'VAF 2026 Q1'),
                    ('hhoc_2025_te', 'HHOC 2025'), ('hhoc_2026_te', 'HHOC 2026 Q1')]:
    d = results[key]
    print(f"\n{'='*60}")
    print(f"{label} Time Entry Summary")
    print(f"{'='*60}")
    print(f"  Entries: {d['total_entries']}, Tickets: {d['unique_tickets']}")
    print(f"  Date Range: {d['date_range'][0]} to {d['date_range'][1]}")
    print(f"  Hours: NH={d['hours']['nh']}, AH={d['hours']['ah']}, Onsite={d['hours']['onsite']}, Total={d['hours']['total']}")
    print(f"  Technicians: {d['tech_count']}")
    print(f"  Proactive: {d['proactive_count']}, Reactive: {d['reactive_count']}")
    print(f"  Weekend entries: {d['weekend_entries']} ({d['weekend_pct']}%)")

if 'vaf_2025_inv' in results:
    d = results['vaf_2025_inv']
    print(f"\n{'='*60}")
    print(f"VAF 2025 Invoice Summary")
    print(f"{'='*60}")
    print(f"  Total Revenue: ${d['total_revenue']:,.2f}")
    print(f"  Monthly Contract: ${d['monthly_total']:,.2f}")
    print(f"  Recurring: ${d['recurring_total']:,.2f}")
    print(f"  WeeklyOut (Overage): ${d['weeklyout_total']:,.2f}")
    print(f"  One-Time: ${d['onetime_total']:,.2f}")
    print(f"  Weekly: ${d['weekly_total']:,.2f}")
    print(f"  By type: {d['by_type']}")
