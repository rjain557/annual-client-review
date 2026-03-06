"""
Comprehensive analysis of BWH Excel files:
  1. allinv_items_too_bwh.xlsx (Invoice items)
  2. all-ticket_timeE_bwh_2025.xlsx (Time entries)
"""

import openpyxl
from collections import Counter, defaultdict
from datetime import datetime

SEPARATOR = "=" * 90
SUB_SEP = "-" * 80


def analyze_sheet_structure(ws, sheet_name):
    """Analyze a single worksheet: headers, row count, sample data."""
    print(f"\n{SUB_SEP}")
    print(f"  SHEET: '{sheet_name}'")
    print(f"{SUB_SEP}")

    # Get all rows as lists
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("  [EMPTY SHEET]")
        return [], []

    headers = list(rows[0])
    data_rows = rows[1:]

    print(f"\n  Row count (excluding header): {len(data_rows)}")
    print(f"  Column count: {len(headers)}")

    print(f"\n  COLUMN HEADERS:")
    for i, h in enumerate(headers, 1):
        print(f"    {i:3d}. {h}")

    # Sample data - first 5 rows
    sample_count = min(5, len(data_rows))
    print(f"\n  SAMPLE DATA (first {sample_count} rows):")
    for row_idx, row in enumerate(data_rows[:sample_count], 1):
        print(f"\n    --- Row {row_idx} ---")
        for col_idx, h in enumerate(headers):
            val = row[col_idx] if col_idx < len(row) else ""
            print(f"      {h}: {val}")

    # Column data types and null counts
    print(f"\n  COLUMN ANALYSIS:")
    print(f"    {'Column':<40} {'Non-Null':>10} {'Null':>8} {'Sample Types':>20}")
    print(f"    {'-'*40} {'-'*10} {'-'*8} {'-'*20}")
    for col_idx, h in enumerate(headers):
        values = [row[col_idx] if col_idx < len(row) else None for row in data_rows]
        non_null = sum(1 for v in values if v is not None and str(v).strip() != "")
        null_count = len(values) - non_null
        types = set()
        for v in values[:50]:  # sample first 50 for types
            if v is not None:
                types.add(type(v).__name__)
        types_str = ", ".join(sorted(types)) if types else "all-null"
        print(f"    {str(h):<40} {non_null:>10} {null_count:>8} {types_str:>20}")

    return headers, data_rows


def analyze_invoice_file(filepath):
    """Deep analysis of the invoice items file."""
    print(f"\n{SEPARATOR}")
    print(f"  FILE 1: INVOICE ITEMS")
    print(f"  Path: {filepath}")
    print(f"{SEPARATOR}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    print(f"\n  Sheet names: {wb.sheetnames}")

    all_headers = {}
    all_data = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        headers, data_rows = analyze_sheet_structure(ws, sn)
        all_headers[sn] = headers
        all_data[sn] = data_rows

    # --- Summary Statistics ---
    print(f"\n{SUB_SEP}")
    print(f"  INVOICE FILE - SUMMARY STATISTICS")
    print(f"{SUB_SEP}")

    for sn in wb.sheetnames:
        headers = all_headers[sn]
        data_rows = all_data[sn]
        if not headers or not data_rows:
            continue

        print(f"\n  Sheet: '{sn}'")
        h_lower = [str(h).lower() if h else "" for h in headers]

        # Helper to find column index by partial name match
        def find_col(keywords, header_list=h_lower):
            for kw in keywords:
                for i, h in enumerate(header_list):
                    if kw in h:
                        return i
            return None

        # Date range
        date_col = find_col(["date", "inv date", "invoice date", "invdate"])
        if date_col is not None:
            dates = []
            for row in data_rows:
                v = row[date_col] if date_col < len(row) else None
                if v is not None:
                    if isinstance(v, datetime):
                        dates.append(v)
                    elif isinstance(v, str):
                        for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"]:
                            try:
                                dates.append(datetime.strptime(v.strip(), fmt))
                                break
                            except:
                                pass
            if dates:
                print(f"\n  Date Range: {min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}")
                print(f"  Total date entries: {len(dates)}")

        # Invoice types
        type_col = find_col(["type", "inv type", "invoice type", "invtype", "category"])
        if type_col is not None:
            types = [str(row[type_col]).strip() for row in data_rows if row[type_col] is not None and str(row[type_col]).strip()]
            type_counts = Counter(types)
            print(f"\n  Unique Invoice Types ({len(type_counts)}):")
            for t, c in type_counts.most_common():
                print(f"    {t:<40} count: {c}")

        # Revenue by invoice type
        amount_col = find_col(["amount", "total", "revenue", "extended", "ext price", "price", "ext_price", "extprice", "line total", "net"])
        if amount_col is not None and type_col is not None:
            revenue_by_type = defaultdict(float)
            total_revenue = 0.0
            for row in data_rows:
                inv_type = str(row[type_col]).strip() if row[type_col] else "Unknown"
                amt = row[amount_col] if amount_col < len(row) else None
                if amt is not None:
                    try:
                        amt_val = float(amt)
                        revenue_by_type[inv_type] += amt_val
                        total_revenue += amt_val
                    except (ValueError, TypeError):
                        pass
            print(f"\n  Total Revenue (all types): ${total_revenue:,.2f}")
            print(f"\n  Revenue by Invoice Type:")
            for t in sorted(revenue_by_type.keys()):
                r = revenue_by_type[t]
                pct = (r / total_revenue * 100) if total_revenue else 0
                print(f"    {t:<40} ${r:>14,.2f}  ({pct:5.1f}%)")
        elif amount_col is not None:
            total_revenue = 0.0
            for row in data_rows:
                amt = row[amount_col] if amount_col < len(row) else None
                if amt is not None:
                    try:
                        total_revenue += float(amt)
                    except (ValueError, TypeError):
                        pass
            print(f"\n  Total Revenue: ${total_revenue:,.2f}")

        # Unique item codes
        item_col = find_col(["item", "item code", "itemcode", "item_code", "product", "sku", "part"])
        if item_col is not None:
            items = set()
            for row in data_rows:
                v = row[item_col] if item_col < len(row) else None
                if v is not None and str(v).strip():
                    items.add(str(v).strip())
            print(f"\n  Unique Item Codes: {len(items)}")
            if len(items) <= 50:
                print(f"  Item codes:")
                for it in sorted(items):
                    count = sum(1 for row in data_rows if row[item_col] is not None and str(row[item_col]).strip() == it)
                    print(f"    {it:<40} count: {count}")
            else:
                print(f"  (Too many to list all - showing top 30 by frequency)")
                item_list = [str(row[item_col]).strip() for row in data_rows if row[item_col] is not None and str(row[item_col]).strip()]
                for it, c in Counter(item_list).most_common(30):
                    print(f"    {it:<40} count: {c}")

        # Unique customers/clients
        client_col = find_col(["client", "customer", "cust", "company", "account"])
        if client_col is not None:
            clients = set()
            for row in data_rows:
                v = row[client_col] if client_col < len(row) else None
                if v is not None and str(v).strip():
                    clients.add(str(v).strip())
            print(f"\n  Unique Clients/Customers: {len(clients)}")
            if len(clients) <= 20:
                for cl in sorted(clients):
                    print(f"    - {cl}")

        # Additional numeric columns - show min/max/avg
        print(f"\n  Numeric Column Statistics:")
        for col_idx, h in enumerate(headers):
            nums = []
            for row in data_rows:
                v = row[col_idx] if col_idx < len(row) else None
                if v is not None:
                    try:
                        nums.append(float(v))
                    except (ValueError, TypeError):
                        pass
            if nums and len(nums) > 5:
                avg = sum(nums) / len(nums)
                print(f"    {h}:")
                print(f"      Min: {min(nums):,.2f}  Max: {max(nums):,.2f}  Avg: {avg:,.2f}  Sum: {sum(nums):,.2f}  Count: {len(nums)}")

    wb.close()


def analyze_time_entries_file(filepath):
    """Deep analysis of the time entries file."""
    print(f"\n\n{SEPARATOR}")
    print(f"  FILE 2: TIME ENTRIES / TICKETS")
    print(f"  Path: {filepath}")
    print(f"{SEPARATOR}")

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    print(f"\n  Sheet names: {wb.sheetnames}")

    all_headers = {}
    all_data = {}

    for sn in wb.sheetnames:
        ws = wb[sn]
        headers, data_rows = analyze_sheet_structure(ws, sn)
        all_headers[sn] = headers
        all_data[sn] = data_rows

    # --- Summary Statistics ---
    print(f"\n{SUB_SEP}")
    print(f"  TIME ENTRIES FILE - SUMMARY STATISTICS")
    print(f"{SUB_SEP}")

    for sn in wb.sheetnames:
        headers = all_headers[sn]
        data_rows = all_data[sn]
        if not headers or not data_rows:
            continue

        print(f"\n  Sheet: '{sn}'")
        h_lower = [str(h).lower() if h else "" for h in headers]

        def find_col(keywords, header_list=h_lower):
            for kw in keywords:
                for i, h in enumerate(header_list):
                    if kw in h:
                        return i
            return None

        # Date range
        date_col = find_col(["date", "entry date", "time date", "work date"])
        if date_col is not None:
            dates = []
            for row in data_rows:
                v = row[date_col] if date_col < len(row) else None
                if v is not None:
                    if isinstance(v, datetime):
                        dates.append(v)
                    elif isinstance(v, str):
                        for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y"]:
                            try:
                                dates.append(datetime.strptime(v.strip(), fmt))
                                break
                            except:
                                pass
            if dates:
                print(f"\n  Date Range: {min(dates).strftime('%Y-%m-%d')} to {max(dates).strftime('%Y-%m-%d')}")
                print(f"  Total date entries: {len(dates)}")

                # Monthly breakdown
                monthly = Counter()
                for d in dates:
                    monthly[d.strftime("%Y-%m")] += 1
                print(f"\n  Monthly Entry Count:")
                for m in sorted(monthly.keys()):
                    print(f"    {m}: {monthly[m]}")

        # Total hours columns - look for NH, AH, Onsite, total hours
        hours_cols = {}
        for i, h in enumerate(h_lower):
            hl = h.strip()
            if any(kw in hl for kw in ["hour", "hrs", "time", " nh", "_nh", "normal", "after", " ah", "_ah", "onsite", "on-site", "on site", "total"]):
                hours_cols[headers[i]] = i
            # Also catch standalone abbreviations
            if hl in ["nh", "ah", "oh", "hours", "hrs", "total hours", "totalhours"]:
                hours_cols[headers[i]] = i

        if hours_cols:
            print(f"\n  Hours-related columns found: {list(hours_cols.keys())}")
            for col_name, col_idx in hours_cols.items():
                total = 0.0
                count = 0
                for row in data_rows:
                    v = row[col_idx] if col_idx < len(row) else None
                    if v is not None:
                        try:
                            fv = float(v)
                            total += fv
                            count += 1
                        except (ValueError, TypeError):
                            pass
                if count > 0:
                    print(f"\n    '{col_name}':")
                    print(f"      Total: {total:,.2f} hours")
                    print(f"      Average per entry: {total/count:,.2f} hours")
                    print(f"      Entries with value: {count}")

        # Role type breakdown
        role_col = find_col(["role", "role type", "roletype", "position", "title", "resource type", "resourcetype"])
        if role_col is not None:
            roles = [str(row[role_col]).strip() for row in data_rows if row[role_col] is not None and str(row[role_col]).strip()]
            role_counts = Counter(roles)
            print(f"\n  Unique Role Types ({len(role_counts)}):")
            for r, c in role_counts.most_common():
                print(f"    {r:<50} entries: {c}")

            # Hours by role
            for col_name, col_idx in hours_cols.items():
                hours_by_role = defaultdict(float)
                for row in data_rows:
                    role = str(row[role_col]).strip() if row[role_col] is not None else "Unknown"
                    v = row[col_idx] if col_idx < len(row) else None
                    if v is not None and role:
                        try:
                            hours_by_role[role] += float(v)
                        except (ValueError, TypeError):
                            pass
                if hours_by_role:
                    total_h = sum(hours_by_role.values())
                    print(f"\n  '{col_name}' by Role Type:")
                    for r in sorted(hours_by_role.keys()):
                        h = hours_by_role[r]
                        pct = (h / total_h * 100) if total_h else 0
                        print(f"    {r:<50} {h:>10,.2f} hrs  ({pct:5.1f}%)")
                    print(f"    {'TOTAL':<50} {total_h:>10,.2f} hrs")

        # Unique technicians
        tech_col = find_col(["tech", "technician", "engineer", "resource", "assigned", "member", "name", "employee", "worker"])
        if tech_col is not None:
            techs = set()
            tech_list = []
            for row in data_rows:
                v = row[tech_col] if tech_col < len(row) else None
                if v is not None and str(v).strip():
                    techs.add(str(v).strip())
                    tech_list.append(str(v).strip())
            print(f"\n  Unique Technicians/Resources: {len(techs)}")
            tech_counts = Counter(tech_list)
            for t, c in tech_counts.most_common():
                print(f"    {t:<50} entries: {c}")

            # Hours by technician
            for col_name, col_idx in hours_cols.items():
                hours_by_tech = defaultdict(float)
                for row in data_rows:
                    tech = str(row[tech_col]).strip() if row[tech_col] is not None else "Unknown"
                    v = row[col_idx] if col_idx < len(row) else None
                    if v is not None and tech:
                        try:
                            hours_by_tech[tech] += float(v)
                        except (ValueError, TypeError):
                            pass
                if hours_by_tech:
                    total_h = sum(hours_by_tech.values())
                    print(f"\n  '{col_name}' by Technician:")
                    for t in sorted(hours_by_tech.keys(), key=lambda x: hours_by_tech[x], reverse=True):
                        h = hours_by_tech[t]
                        pct = (h / total_h * 100) if total_h else 0
                        print(f"    {t:<50} {h:>10,.2f} hrs  ({pct:5.1f}%)")
                    print(f"    {'TOTAL':<50} {total_h:>10,.2f} hrs")

        # Unique tickets
        ticket_col = find_col(["ticket", "ticket #", "ticket#", "ticketid", "ticket_id", "ticket number", "ticketnumber", "case", "incident", "sr #", "sr#"])
        if ticket_col is not None:
            tickets = set()
            for row in data_rows:
                v = row[ticket_col] if ticket_col < len(row) else None
                if v is not None and str(v).strip():
                    tickets.add(str(v).strip())
            print(f"\n  Unique Tickets: {len(tickets)}")

        # Additional numeric columns stats
        print(f"\n  Numeric Column Statistics:")
        for col_idx, h in enumerate(headers):
            nums = []
            for row in data_rows:
                v = row[col_idx] if col_idx < len(row) else None
                if v is not None:
                    try:
                        nums.append(float(v))
                    except (ValueError, TypeError):
                        pass
            if nums and len(nums) > 5:
                avg = sum(nums) / len(nums)
                print(f"    {h}:")
                print(f"      Min: {min(nums):,.4f}  Max: {max(nums):,.4f}  Avg: {avg:,.4f}  Sum: {sum(nums):,.2f}  Count: {len(nums)}")

        # Unique values for key categorical columns
        print(f"\n  Unique Value Counts for All Columns:")
        for col_idx, h in enumerate(headers):
            vals = set()
            for row in data_rows:
                v = row[col_idx] if col_idx < len(row) else None
                if v is not None and str(v).strip():
                    vals.add(str(v).strip())
            print(f"    {h:<40} {len(vals):>6} unique values")
            # If <= 15 unique values, list them
            if 1 < len(vals) <= 15:
                val_list = [str(row[col_idx]).strip() for row in data_rows if row[col_idx] is not None and str(row[col_idx]).strip()]
                val_counts = Counter(val_list)
                for v, c in val_counts.most_common():
                    print(f"      - {v:<36} ({c})")

    wb.close()


if __name__ == "__main__":
    print("\n" + "=" * 90)
    print("  BWH CLIENT - COMPREHENSIVE EXCEL FILE ANALYSIS")
    print("  Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    print("=" * 90)

    invoice_file = r"c:\vscode\annual-client-review\annual-client-review\clients\bwh\2025\allinv_items_too_bwh.xlsx"
    time_file = r"c:\vscode\annual-client-review\annual-client-review\clients\bwh\2025\all-ticket_timeE_bwh_2025.xlsx"

    analyze_invoice_file(invoice_file)
    analyze_time_entries_file(time_file)

    print(f"\n\n{'=' * 90}")
    print("  ANALYSIS COMPLETE")
    print(f"{'=' * 90}\n")
