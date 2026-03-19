"""Full HHOC 2025 invoice analysis from corrected file."""
import openpyxl
from collections import defaultdict
from datetime import datetime
import json, os

BASE = r"c:\vscode\annual-client-review\annual-client-review"

def safe_float(v):
    if v is None or v == 'NULL' or v == '':
        return 0.0
    try:
        return float(v)
    except:
        return 0.0

wb = openpyxl.load_workbook(os.path.join(BASE, 'clients/hhoc/2025/hhoc_inv_from_2025_corrected.xlsx'), read_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
data = rows[1:]
wb.close()

print(f"Total rows: {len(data)}")
print(f"Headers: {headers}")

# Filter to 2025
inv_2025 = [r for r in data if r[1] and isinstance(r[1], datetime) and r[1].year == 2025]
inv_2026 = [r for r in data if r[1] and isinstance(r[1], datetime) and r[1].year == 2026]
print(f"2025 rows: {len(inv_2025)}")
print(f"2026 rows: {len(inv_2026)}")

# By invoice type
print("\n=== REVENUE BY INVOICE TYPE (2025) ===")
type_rev = defaultdict(float)
type_count = defaultdict(int)
for r in inv_2025:
    itype = str(r[2]) if r[2] else 'Unknown'
    type_rev[itype] += safe_float(r[9])
    type_count[itype] += 1

total_rev = 0
for t in sorted(type_rev.keys(), key=lambda x: -type_rev[x]):
    print(f"  {t:15s} ${type_rev[t]:>12,.2f}  ({type_count[t]} line items)")
    total_rev += type_rev[t]
print(f"  {'TOTAL':15s} ${total_rev:>12,.2f}")

# All unique items
print("\n=== ALL UNIQUE ITEMS ===")
items_set = set()
for r in inv_2025:
    item = str(r[3]) if r[3] else ''
    qb = str(r[8]) if r[8] else ''
    itype = str(r[2]) if r[2] else ''
    items_set.add((item, qb, itype))
for item, qb, it in sorted(items_set):
    print(f"  [{it:12s}] {item:40s} | {qb}")

# Monthly breakdown by category
print("\n=== MONTHLY CONTRACT BY CATEGORY (2025) ===")
monthly_items = [r for r in inv_2025 if str(r[2]) == 'Monthly']
monthly_cats = defaultdict(float)
for r in monthly_items:
    item = str(r[3]) if r[3] else ''
    qb = str(r[8]) if r[8] else ''
    total = safe_float(r[9])

    if 'Labor' in qb or item.startswith('OffShore') or item.startswith('Tech_Support'):
        monthly_cats['Labor'] += total
    elif any(x in qb for x in ['Crowdstrike', 'Huntress', 'Malware']) or item in ['AVD','AVS','AVHS','AVMD','AVMS','AVMH']:
        monthly_cats['Security & Endpoint'] += total
    elif 'Backup' in qb or item in ['IB', 'VONE', 'V365']:
        monthly_cats['Backup & Archiving'] += total
    elif any(x in qb for x in ['RMM', 'My Remote', 'My Ops', 'MyDisk']) or item in ['PMW','MR','OPS-BKP','OPS-NET','OPS-PRT','OPS-WF','OPS-TR','OPS-ST','MDU']:
        monthly_cats['Monitoring & Management'] += total
    elif 'SIP' in qb or item in ['SIP','DID','TFN','SMS','Incoming']:
        monthly_cats['VoIP / Telephony'] += total
    elif any(x in qb for x in ['VPS', 'Cloud']) or item in ['CL-GB','CL-VC','TB-BSTR','TB-PSTR','TB-RSTR']:
        monthly_cats['Cloud Hosting'] += total
    elif 'Secure Internet' in qb or item == 'SI':
        monthly_cats['Secure Internet'] += total
    elif 'PenTesting' in qb or item in ['PT','RTPT']:
        monthly_cats['Pen Testing'] += total
    elif 'Assessment' in qb or item in ['NA','SA']:
        monthly_cats['Assessment'] += total
    elif any(x in qb for x in ['Sophos', 'Edge', 'Velocloud']) or item in ['SO-1C4G','SO-2C4G','Edge-16M','VC-100M']:
        monthly_cats['Network & Firewall'] += total
    elif 'Anti-Spam' in qb or 'DMARC' in qb or 'DKIM' in qb or item in ['ASA','DKIM','PHT']:
        monthly_cats['Email Security'] += total
    elif 'Cyber' in qb or item == 'CT':
        monthly_cats['Cybertraining'] += total
    else:
        monthly_cats[f'Other ({item})'] += total

for cat in sorted(monthly_cats.keys(), key=lambda x: -monthly_cats[x]):
    v = monthly_cats[cat]
    print(f"  {cat:35s} ${v:>10,.2f}  (~${v/12:>8,.2f}/mo)")
monthly_total = sum(monthly_cats.values())
print(f"  {'TOTAL':35s} ${monthly_total:>10,.2f}  (~${monthly_total/12:>8,.2f}/mo)")

# Monthly run rate
print("\n=== MONTHLY RUN RATE ===")
monthly_by_month = defaultdict(float)
for r in inv_2025:
    if str(r[2]) == 'Monthly' and r[1] and isinstance(r[1], datetime):
        monthly_by_month[r[1].strftime('%Y-%m')] += safe_float(r[9])
for m in sorted(monthly_by_month.keys()):
    print(f"  {m}: ${monthly_by_month[m]:>10,.2f}")

# Labor detail by month
print("\n=== LABOR BY MONTH ===")
labor_items = [r for r in inv_2025 if str(r[2]) == 'Monthly' and
               ('Labor' in str(r[8]) or str(r[3]).startswith('OffShore') or str(r[3]).startswith('Tech_Support'))]
labor_by_month = defaultdict(list)
for r in labor_items:
    if r[1] and isinstance(r[1], datetime):
        labor_by_month[r[1].strftime('%Y-%m')].append({
            'item': str(r[3]),
            'desc': str(r[4])[:50] if r[4] else '',
            'qty': safe_float(r[6]),
            'price': safe_float(r[7]),
            'total': safe_float(r[9])
        })
for m in sorted(labor_by_month.keys()):
    items = labor_by_month[m]
    total = sum(i['total'] for i in items)
    print(f"  {m}: ${total:>8,.2f}")
    for i in items:
        print(f"    {i['item']:30s} Qty={i['qty']:>6} Rate=${i['price']:>8.2f} Total=${i['total']:>8,.2f}")

# Recurring items
print("\n=== RECURRING ITEMS (2025) ===")
recurring = [r for r in inv_2025 if str(r[2]) == 'Recurring']
rec_items = defaultdict(float)
for r in recurring:
    desc = str(r[4]) if r[4] else str(r[3])
    rec_items[desc] += safe_float(r[9])
rec_total = 0
for k, v in sorted(rec_items.items(), key=lambda x: -x[1]):
    print(f"  {k:50s} ${v:>10,.2f}")
    rec_total += v
print(f"  {'TOTAL':50s} ${rec_total:>10,.2f}")

# WeeklyOut (overage)
print("\n=== WEEKLYOUT (LABOR OVERAGE) ===")
weeklyout = [r for r in inv_2025 if str(r[2]) == 'WeeklyOut']
wo_total = 0
for r in weeklyout:
    total = safe_float(r[9])
    wo_total += total
    print(f"  {r[1]} {str(r[3]):30s} Qty={r[6]} Rate=${safe_float(r[7]):.0f} Total=${total:,.2f}")
print(f"  TOTAL: ${wo_total:,.2f}")

# One-Time / Invoice
print("\n=== ONE-TIME / INVOICE ITEMS ===")
onetime = [r for r in inv_2025 if str(r[2]) in ('One-Time', 'Invoice')]
ot_total = 0
for r in sorted(onetime, key=lambda x: x[1] if x[1] else datetime(2025,1,1)):
    total = safe_float(r[9])
    ot_total += total
    date = r[1].strftime('%Y-%m-%d') if r[1] and isinstance(r[1], datetime) else '?'
    print(f"  {date} {str(r[3]):30s} {str(r[4])[:50]:50s} ${total:>10,.2f}")
print(f"  TOTAL: ${ot_total:,.2f}")

# Weekly time tracking
print("\n=== WEEKLY (LABOR TRACKING) ===")
weekly = [r for r in inv_2025 if str(r[2]) == 'Weekly']
weekly_total = sum(safe_float(r[9]) for r in weekly)
print(f"  {len(weekly)} line items, Total: ${weekly_total:,.2f}")

# Latest month invoice detail
print("\n=== LATEST MONTH INVOICE DETAIL ===")
if monthly_by_month:
    latest = max(monthly_by_month.keys())
    latest_items = [r for r in inv_2025 if str(r[2]) == 'Monthly' and r[1] and
                   isinstance(r[1], datetime) and r[1].strftime('%Y-%m') == latest]
    print(f"Month: {latest}")
    for r in sorted(latest_items, key=lambda x: -safe_float(x[9])):
        print(f"  {str(r[3]):20s} {str(r[4])[:45]:45s} Qty={safe_float(r[6]):>6} Rate=${safe_float(r[7]):>8.2f} Total=${safe_float(r[9]):>8,.2f}")

# Grand total
print(f"\n=== HHOC 2025 GRAND TOTAL ===")
print(f"  Monthly Contract: ${monthly_total:>12,.2f}")
print(f"  Recurring:        ${rec_total:>12,.2f}")
print(f"  WeeklyOut:        ${wo_total:>12,.2f}")
print(f"  One-Time/Invoice: ${ot_total:>12,.2f}")
print(f"  Weekly:           ${weekly_total:>12,.2f}")
print(f"  TOTAL:            ${total_rev:>12,.2f}")
