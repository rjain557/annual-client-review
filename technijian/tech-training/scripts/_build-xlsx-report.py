"""Build Technijian-branded Excel report for the tech-time-entry audit.

Reads the three outlier CSVs produced by _flag-outliers.py and emits a
multi-tab, styled xlsx at:
  technijian/tech-training/<client>/<year>/<CLIENT>-Tech-Time-Entry-Audit.xlsx

Usage: python _build-xlsx-report.py <client_code> <year>
"""
import csv
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

SCRIPTS = Path(__file__).resolve().parent
REPO = SCRIPTS.parent.parent.parent
CLIENT = (sys.argv[1] if len(sys.argv) > 1 else "bwh").lower()
YEAR = sys.argv[2] if len(sys.argv) > 2 else "2026"
BASE = REPO / "technijian" / "tech-training" / CLIENT / YEAR
OUT = BASE / f"{CLIENT.upper()}-Tech-Time-Entry-Audit.xlsx"

# --- brand colors (from Technijian Brand Guide 2026) ---
CORE_BLUE = "006DB6"
CORE_ORANGE = "F67D4B"
TEAL = "1EAAC8"
DARK_CHARCOAL = "1A1A2E"
BRAND_GREY = "59595B"
OFF_WHITE = "F8F9FA"
LIGHT_GREY = "E9ECEF"
WHITE = "FFFFFF"
GREEN = "28A745"
RED = "CC0000"

FONT = "Open Sans"

HEADER_FILL = PatternFill("solid", fgColor=CORE_BLUE)
HEADER_FONT = Font(name=FONT, color=WHITE, bold=True, size=11)
TITLE_FONT = Font(name=FONT, color=CORE_BLUE, bold=True, size=18)
SUBTITLE_FONT = Font(name=FONT, color=BRAND_GREY, size=11)
LABEL_FONT = Font(name=FONT, color=BRAND_GREY, size=10)
VALUE_FONT = Font(name=FONT, color=CORE_BLUE, bold=True, size=20)
BODY_FONT = Font(name=FONT, color=BRAND_GREY, size=10)
TOTAL_FILL = PatternFill("solid", fgColor=LIGHT_GREY)
TOTAL_FONT = Font(name=FONT, color=DARK_CHARCOAL, bold=True, size=10)
KPI_FILL = PatternFill("solid", fgColor=OFF_WHITE)
THIN = Side(style="thin", color=LIGHT_GREY)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def read_csv(path: Path):
    with path.open("r", encoding="utf-8", newline="") as f:
        rows = list(csv.reader(f))
    return rows[0], rows[1:]


def style_header_row(ws, row_idx, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER


def style_body_row(ws, row_idx, n_cols, zebra=False):
    fill = PatternFill("solid", fgColor=OFF_WHITE) if zebra else None
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.font = BODY_FONT
        cell.border = BORDER
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --- Build workbook ---
wb = openpyxl.Workbook()

# ============ Sheet 1: Summary ============
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:F1")
ws["A1"] = f"Tech Time-Entry Audit — {CLIENT.upper()} {YEAR}"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = Alignment(vertical="center")
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:F2")
ws["A2"] = "Flags individual time entries where hours claimed appear outside reasonable range for the work described"
ws["A2"].font = SUBTITLE_FONT

# KPI row (row 5-6)
# Use 4 KPI cards
bytech_hdr, bytech_rows = read_csv(BASE / "tech-outliers-by-tech.csv")
total_entries = sum(int(r[1]) for r in bytech_rows)
total_hours = sum(float(r[2]) for r in bytech_rows)
flagged_entries = sum(int(r[3]) for r in bytech_rows)
flagged_hours = sum(float(r[4]) for r in bytech_rows)

# Re-read with full flagged-detail count
# (bytech only shows techs with flags; total comes from per-entry CSV)
detail_hdr, detail_rows = read_csv(BASE / "tech-outliers-detail.csv")

# For "total entries" we want the SCANNED number, not flagged sum.
# Re-derive by loading time-entries.csv directly.
te_csv = REPO / "clients" / CLIENT / YEAR / "03_Accounting" / "time-entries.csv"
scanned_entries = 0
scanned_hours = 0.0
with te_csv.open("r", encoding="utf-8", newline="") as f:
    reader = csv.DictReader(f)
    for row in reader:
        try:
            h = abs(float(row.get("Hours") or 0))
        except ValueError:
            h = 0.0
        if h > 0:
            scanned_entries += 1
            scanned_hours += h

kpi_row = 5
kpi_cells = [
    ("A", f"{scanned_entries:,}", "Total Entries Scanned"),
    ("B", f"{scanned_hours:,.0f}", "Total Hours Scanned"),
    ("D", f"{flagged_entries}", "Flagged Entries"),
    ("E", f"{flagged_hours:,.1f}", "Flagged Hours"),
    ("F", f"{flagged_hours/scanned_hours*100:.1f}%", "% of Hours Flagged"),
]
for col, value, label in kpi_cells:
    vcell = ws[f"{col}{kpi_row}"]
    vcell.value = value
    vcell.font = VALUE_FONT
    vcell.fill = KPI_FILL
    vcell.alignment = Alignment(horizontal="center", vertical="center")
    vcell.border = BORDER
    lcell = ws[f"{col}{kpi_row+1}"]
    lcell.value = label
    lcell.font = LABEL_FONT
    lcell.fill = KPI_FILL
    lcell.alignment = Alignment(horizontal="center", vertical="center")
    lcell.border = BORDER
ws.row_dimensions[kpi_row].height = 36

# Flag legend
r = 9
ws.cell(row=r, column=1, value="Flag Code").font = HEADER_FONT
ws.cell(row=r, column=1).fill = HEADER_FILL
ws.cell(row=r, column=2, value="Meaning").font = HEADER_FONT
ws.cell(row=r, column=2).fill = HEADER_FILL
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
legend = [
    ("H1", "Routine/low-complexity work with hours exceeding the reasonable category cap"),
    ("H2", "Generic/vague title (\"Help\", \"Test\", \"Fix\") with more than 0.5 hour claimed"),
    ("H3", "Single entry > 8 hours (suggests a whole-day dump into one block)"),
    ("H4", "Tech's total on one date > 12 hours (cross-ticket over-claim)"),
    ("H5", "Same tech + ticket + day with multiple entries totalling > 2 × cap"),
]
for i, (code, desc) in enumerate(legend, 1):
    ws.cell(row=r+i, column=1, value=code).font = Font(name=FONT, bold=True, color=CORE_ORANGE, size=10)
    ws.cell(row=r+i, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=r+i, column=1).border = BORDER
    ws.cell(row=r+i, column=2, value=desc).font = BODY_FONT
    ws.cell(row=r+i, column=2).border = BORDER
    ws.merge_cells(start_row=r+i, start_column=2, end_row=r+i, end_column=6)

set_col_widths(ws, [16, 16, 16, 16, 16, 26])

# ============ Sheet 2: Techs Ranked ============
ws2 = wb.create_sheet("Techs Ranked")
ws2.sheet_view.showGridLines = False
ws2.merge_cells("A1:K1")
ws2["A1"] = "Techs Ranked by Flagged Hours"
ws2["A1"].font = TITLE_FONT
ws2.row_dimensions[1].height = 28

# header
for i, h in enumerate(bytech_hdr, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, len(bytech_hdr))
ws2.row_dimensions[3].height = 26

# body
for i, row in enumerate(bytech_rows, 4):
    for j, v in enumerate(row, 1):
        # Cast numerics for sorting
        cell = ws2.cell(row=i, column=j)
        if j in (2, 3, 4, 5):
            try:
                cell.value = float(v) if "." in v else int(v)
            except ValueError:
                cell.value = v
        elif j == 6:
            # FlaggedPct string like "13.1%"
            try:
                cell.value = float(v.rstrip("%")) / 100
                cell.number_format = "0.0%"
            except ValueError:
                cell.value = v
        elif j >= 7:
            try:
                cell.value = int(v)
            except ValueError:
                cell.value = v
        else:
            cell.value = v
    style_body_row(ws2, i, len(bytech_hdr), zebra=(i % 2 == 0))

# format hour cols
for col in ("C", "E"):
    for cell in ws2[col][3:]:
        cell.number_format = "#,##0.0"

# conditional format flagged % column (F)
last_row = 3 + len(bytech_rows)
ws2.conditional_formatting.add(
    f"F4:F{last_row}",
    CellIsRule(operator="greaterThanOrEqual", formula=["0.15"],
               fill=PatternFill("solid", fgColor="FCDADA")),
)
ws2.conditional_formatting.add(
    f"F4:F{last_row}",
    CellIsRule(operator="between", formula=["0.05", "0.1499"],
               fill=PatternFill("solid", fgColor="FFF1DF")),
)

set_col_widths(ws2, [18, 14, 14, 14, 14, 14, 8, 8, 8, 8, 8])
ws2.freeze_panes = "A4"
ws2.auto_filter.ref = f"A3:K{last_row}"

# ============ Sheet 3: Flagged Entries Detail ============
ws3 = wb.create_sheet("Flagged Entries Detail")
ws3.sheet_view.showGridLines = False
ws3.merge_cells("A1:L1")
ws3["A1"] = f"All Flagged Entries ({len(detail_rows)} rows)"
ws3["A1"].font = TITLE_FONT
ws3.row_dimensions[1].height = 28

for i, h in enumerate(detail_hdr, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, len(detail_hdr))
ws3.row_dimensions[3].height = 26

# Cast Hours, CategoryCap, DailyTotal as numbers
hours_col = detail_hdr.index("Hours") + 1
cap_col = detail_hdr.index("CategoryCap") + 1
day_col = detail_hdr.index("DailyTotal") + 1
numeric_cols = {hours_col, cap_col, day_col}

for i, row in enumerate(detail_rows, 4):
    for j, v in enumerate(row, 1):
        cell = ws3.cell(row=i, column=j)
        if j in numeric_cols:
            try:
                cell.value = float(v)
                cell.number_format = "0.00"
            except ValueError:
                cell.value = v
        else:
            cell.value = v
    style_body_row(ws3, i, len(detail_hdr), zebra=(i % 2 == 0))

# Colour-code Hours column vs Cap
last = 3 + len(detail_rows)
ws3.conditional_formatting.add(
    f"{get_column_letter(hours_col)}4:{get_column_letter(hours_col)}{last}",
    ColorScaleRule(start_type="min", start_color="FFF1DF",
                   mid_type="percentile", mid_value=50, mid_color="FBBF77",
                   end_type="max", end_color="E85A3A"),
)

set_col_widths(ws3, [11, 18, 10, 8, 48, 30, 10, 10, 12, 10, 46, 22])
ws3.freeze_panes = "A4"
ws3.auto_filter.ref = f"A3:{get_column_letter(len(detail_hdr))}{last}"

# ============ Sheet 4: Category Caps (the rules used) ============
ws4 = wb.create_sheet("Category Caps")
ws4.sheet_view.showGridLines = False
ws4.merge_cells("A1:C1")
ws4["A1"] = "Category Caps (single-entry hour limit above which an entry is flagged H1)"
ws4["A1"].font = TITLE_FONT
ws4.row_dimensions[1].height = 28

# Pull cap dict by reloading the flag-outliers module
import importlib.util as _ilu
_fo_spec = _ilu.spec_from_file_location("fo", SCRIPTS / "_flag-outliers.py")
# We can't exec the module easily because it runs the analysis. Instead
# read the caps by parsing the CATEGORY_CAP source directly.
fo_src = (SCRIPTS / "_flag-outliers.py").read_text(encoding="utf-8")
import re as _re
cap_rows = _re.findall(r'"(Routine: .+?|Project: .+?|Uncategorized)":\s*([\d.]+)', fo_src)
ws4.cell(row=3, column=1, value="Category").fill = HEADER_FILL
ws4.cell(row=3, column=1).font = HEADER_FONT
ws4.cell(row=3, column=2, value="Single-entry cap (hours)").fill = HEADER_FILL
ws4.cell(row=3, column=2).font = HEADER_FONT
ws4.cell(row=3, column=3, value="Type").fill = HEADER_FILL
ws4.cell(row=3, column=3).font = HEADER_FONT
for i, (cat, cap) in enumerate(cap_rows, 4):
    ws4.cell(row=i, column=1, value=cat).font = BODY_FONT
    ws4.cell(row=i, column=2, value=float(cap)).font = BODY_FONT
    ws4.cell(row=i, column=2).number_format = "0.00"
    if cat.startswith("Project:"):
        kind = "Project (higher cap)"
        color = CORE_ORANGE
    elif cat.startswith("Routine:"):
        kind = "Routine"
        color = TEAL
    else:
        kind = "Fallback"
        color = BRAND_GREY
    ws4.cell(row=i, column=3, value=kind).font = Font(name=FONT, color=color, size=10, bold=True)
    for c in (1, 2, 3):
        ws4.cell(row=i, column=c).border = BORDER
        if i % 2 == 0:
            ws4.cell(row=i, column=c).fill = PatternFill("solid", fgColor=OFF_WHITE)

set_col_widths(ws4, [60, 22, 22])
ws4.freeze_panes = "A4"

# ============ Sheet 5: Top 25 Worst Entries ============
ws5 = wb.create_sheet("Top 25 Worst Entries")
ws5.sheet_view.showGridLines = False
ws5.merge_cells("A1:F1")
ws5["A1"] = "Top 25 Worst Flagged Entries (by hours)"
ws5["A1"].font = TITLE_FONT
ws5.row_dimensions[1].height = 28

# Sort detail_rows by hours desc, take top 25
top25 = sorted(detail_rows, key=lambda r: -float(r[hours_col - 1]))[:25]
top_hdr = ["Date", "Tech", "Hours", "Cap", "Title", "Flags"]
for i, h in enumerate(top_hdr, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, len(top_hdr))
ws5.row_dimensions[3].height = 26

for i, row in enumerate(top25, 4):
    ws5.cell(row=i, column=1, value=row[0])
    ws5.cell(row=i, column=2, value=row[1])
    ws5.cell(row=i, column=3, value=float(row[hours_col - 1]))
    ws5.cell(row=i, column=3).number_format = "0.00"
    ws5.cell(row=i, column=4, value=float(row[cap_col - 1]))
    ws5.cell(row=i, column=4).number_format = "0.00"
    ws5.cell(row=i, column=5, value=row[4])  # Title
    ws5.cell(row=i, column=6, value=row[detail_hdr.index("Flags")])
    style_body_row(ws5, i, len(top_hdr), zebra=(i % 2 == 0))

set_col_widths(ws5, [11, 18, 10, 10, 70, 12])
ws5.freeze_panes = "A4"

# Save
wb.save(str(OUT))
print(f"Wrote: {OUT}")
